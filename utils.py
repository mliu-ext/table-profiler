#!/usr/bin/env python
# coding: utf-8

# functions and tools used for comparing and profiling tables in snowflake

import sys
import json
import pandas as pd
import snowflake.connector
from time import gmtime, strftime
from tqdm.notebook import tqdm

from profile_query import *


# list of numeric data types in snowflake tables
SN_NUMERIC = ['NUMBER', 'DECIMAL', 'NUMERIC', 'INT', 'INTEGER', 'BIGINT', 'SMALLINT','FLOAT', 'FLOAT4', 'FLOAT8', 'DOUBLE', 'DOUBLE PRECISION', 'REAL']   

def sn_conn(SNOW_CONFIG):
    con = snowflake.connector.connect(
        user = SNOW_CONFIG['user'],
        account = SNOW_CONFIG['account'],
        authenticator = 'externalbrowser',
        role = SNOW_CONFIG['role'],
        warehouse = SNOW_CONFIG['warehouse']
        )
    return con

def append_excel(dataframe, filename, sheetname='Sheet1', index=True):
    """
    append dataframe to existing excel file
    dataframe is pandas dataframe
    filename is a string of excel file name including the path
    sheetname is excel sheetname
    Set index True to write df index to file
    """
    import os.path
    from openpyxl import load_workbook
    import pandas as pd
    if os.path.exists(filename):
        book = load_workbook(filename)
        if sheetname not in [ws.title for ws in book.worksheets]:
            book.create_sheet(sheetname)
            header = True
            # print(sheetname)
        else:
            header = False
        writer = pd.ExcelWriter(filename, engine='openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        # print(writer.sheets)
        dataframe.to_excel(writer, sheet_name=sheetname, startrow=writer.book[sheetname].max_row, header=header, index=index)
        writer.save()
    else:
        with pd.ExcelWriter(filename) as writer: 
            dataframe.to_excel(writer, sheet_name=sheetname, index=index, header=True)


class TableStats:

    def __init__(self, table, con, engine='snowflake', where_condition=''):
        self.table = table
        self.con = con
        self.engine = engine
        self.check_table_exist()
        self.where_condition = where_condition
        self.columns = self.get_column_names()
        self.row_count = self.get_row_count()
        self.col_count = len(self.columns) if self.columns is not None else 0
        self.get_partition_name_and_value = self.get_partition_name_and_value

    def add_where_condition(self):
        if (self.where_condition!="") & (~self.where_condition.isspace()):
            where_condition = 'where ' + self.where_condition
        else:
            where_condition = ''
        return where_condition

    def all_condition(self, cond):
        """
        add additional conditon to where condition
        :return: a string
        """
        if ~cond.isspace() & ('where' in self.add_where_condition()):
            return self.add_where_condition() + ' and ' + cond
        elif ~cond.isspace() & (self.add_where_condition() == ''):
            return 'where ' + cond
        else:
            return self.add_where_condition()

    def check_table_exist(self):
        try:
            self.snsql("desc table {}".format(self.table))
        except:
            print(sys.exc_info())
            sys.exit("could NOT describe table {}. Check if table name is correct and in the format: db_name.schema_name.table_name".format(self.table))
        
    def snsql(self, query):
        df = pd.read_sql(query, self.con)
        df.columns = [s.lower() for s in df.columns]
        return df

    def add_lower(self, column):
        # make column lower case
        cols_list = column.split(',')
        cols_lower = ','.join(['lower({c}) {c}'.format(c=col.strip()) for col in cols_list])
        return cols_lower

    def db(self):
        """
        find the right engine to run query
        """
        if self.engine == 'snowflake':
            return self.snsql
        else:
            raise NameError("engine name has to be in the list [snowflake,]")
            sys.exit(1)

    def get_partition_info(self):
        """
        run a query to fetch partition name and type
        :return: dataframe
        """
        query = 'desc {}'.format(self.table)

        try:
            res = self.db()(query)
        except:
            sys.exit(sys.exc_info())

        if self.engine.lower() == 'presto':
            df = res[res.iloc[:, 2] == 'partition key'].iloc[:, :2]
            df.columns = ['partition', 'type']
            if len(df.index) == 0:
                df = pd.DataFrame({"partition": 'no partition', 'type': 'n/a'}, index=[0])
        else:
            df = pd.DataFrame({"partition": 'can only find partition in presto engine', 'type': 'n/a'}, index=[0])
        # find the latest and earliest partition if no partition set the value to 'n/a'
        df = df.reset_index(drop=True)
        # print("partition type is: {}".format(df.type[0]))
        if df.type[0] != 'n/a':
            sql = "select max({}), min({}) from {}".format(df.partition[0], df.partition[0], self.table)
            try:
                latest = self.db()(sql)
                df['latest'] = latest.iloc[0, 0]
                df['earliest'] = latest.iloc[0, 1]
            except:
                sys.exit(sys.exc_info())
        else:
            df['latest'] = 'n/a'
            df['earliest'] = 'n/a'
        # print(df)
        return df


    def get_partition_name_and_value(self):
        [partition_name, partition_value] = map(lambda x: x.strip(), self.where_condition.split('='))
        return partition_name, partition_value


    def get_row_count(self):
        """
        Count number of rows in the table
        """
        query = '''select count(1) from {0}  {1}'''.format(self.table, self.add_where_condition())
        try:
            return self.db()(query).iloc[0, 0]
        except:
            sys.exit(sys.exc_info())


    def get_column_names(self):
        """
        Get column names of table
        """
        query = """ show columns in {t}""".format(t=self.table)
        try:
            df = self.db()(query).loc[:, 'column_name']
            df = pd.DataFrame(df)
            df.columns = ['col_name']
            return df
        except:
            sys.exit(sys.exc_info())


    def get_column_names_with_type(self):
        """
        Get column names and types of table
        """
        query = """ show columns in {t}""".format(t=self.table)
        try:
            df = self.db()(query).iloc[:, [2, 3]]
            df.columns = ['col_name', 'type']
            df.loc[:, 'type'] = df['type'].apply(lambda s: json.loads(s)['type'])
            return df
        except:
            sys.exit(sys.exc_info())


    def get_distinct_value(self, column):
        """
        query distinct value of one or more columns
        """
        query = "select distinct {c} from {t} {f} order by {c}".format(t=self.table, col=self.add_lower(column), c=column, f=self.add_where_condition())
        # print(query)
        try:
            return self.db()(query)
        except:
            sys.exit(sys.exc_info())


    def get_distinct_count(self, column):
        """
        query number of distinct values in one or more columns
        :param column: a string with column name use comma to separate column names
        :return: DataFrame
        """
        query = """ select count(distinct {c}) from {t} {f} """.format(c=column, t=self.table, f=self.add_where_condition())
        # print(query)
        try:
            return self.db()(query)
        except:
            sys.exit(sys.exc_info())


    def get_group_count_lower_case(self, column, limit=1000, cond=' '):
        """
        query count group by a column;
        return dataframe with distinct value in the column and the corresponding count 
        """
        all_cond = self.all_condition(cond)
        cols_lower = self.add_lower(column)
        query = """ select {lowerC}, count(1) cnt from {t} {f} group by {c} order by cnt desc, {c} limit {limit}
                 """.format(lowerC=cols_lower, c=column, t=self.table, f=all_cond, limit=limit)
        # print(query)
        try:
            return self.db()(query)
        except:
            sys.exit(sys.exc_info())


    def get_group_count_with_case(self, column, limit=1000, cond=' '):
        """
        query count group by a column;
        return dataframe with distinct value in the column and the corresponding count
        """
        all_cond = self.all_condition(cond)
        query = """ select {c}, count(1) cnt from {t} {f} group by {c} order by cnt desc, {c} limit {limit}
                 """.format(c=column, t=self.table, f=all_cond, limit=limit)
        # print(query)
        try:
            return self.db()(query)
        except:
            sys.exit(sys.exc_info())


    def get_group_count(self, column, limit=1000, cond=' '):
        """
        query count group by a column;
        return dataframe with distinct value in the column and the corresponding count
        """
        all_cond = self.all_condition(cond)
        query = """ select {c}, count(1) cnt from {t} {f} group by {c} order by cnt desc, {c} limit {limit}
                 """.format(c=column, t=self.table, f=all_cond, limit=limit)
        # print(query)
        try:
            return self.db()(query)
        except:
            sys.exit(sys.exc_info())


    def get_group_agg(self, column, agg='count(1)', limit=1000, cond=' '):
        """
        :param column: one or more column names to group by
        :param agg: aggregation string
        :param limit: number of rows to output
        :param cond: additional conditions in where clause
        :return: DataFrame with distinct value in the column and the corresponding count
        """
        all_cond = self.all_condition(cond)
        query = """ select {c}, {agg} agg from {t} {f} group by {c} order by agg desc, {c} limit {limit}
                 """.format(c=column, agg=agg, t=self.table, f=all_cond, limit=limit)
        # print(query)
        try:
            return self.db()(query)
        except:
            sys.exit(sys.exc_info())


    def profile_all_columns(self, filename=''):
        # profile each column of the table 
        # Different metrics were calculated based on column data types
        # Numeric columns: RowCount, NullCount, Duplicatecount, Average, Max, Min, BlankCount
        # non Numeric columns: RowCount, DistinctCount, NullCount, UniqueCount, Duplicatecount, BlankCount

        columns = self.get_column_names_with_type()
        print()
        print(columns)
        print()
        col_list = columns.col_name.to_list()  

        for col in col_list:
            col = col.strip()
            print("profiling column: {}".format(col))
            col_type = columns[columns.col_name == col].type.iloc[0]
            
            if  col_type in SN_NUMERIC:  # run query for numeric columns
                sql_t = sql_num
            else:  # run query for varchar columns
                sql_t = sql_char
                   
            # run all queries to get summary of the column
            dft = self.snsql(sql_t.format(column=col, table=self.table))
            dfa = self.snsql(sqla.format(column=col, table=self.table))        
            dfb = self.snsql(sqlb.format(column=col, table=self.table))        
            
            # join the summary together
            dft = pd.DataFrame([{'col_name': col}]) \
                    .merge(dfa, left_index=True, right_index=True) \
                    .merge(dfb, left_index=True, right_index=True) \
                    .merge(dft, left_index=True, right_index=True)
            
            if col == col_list[0]:
                df = dft.copy()
            else:
                df = df.append(dft, sort=False)

        print(df)

        if filename.strip():
            print("Saving table profiling results to {}".format(filename))    
            append_excel(df, filename, sheetname='Sheet1', index=False)

        return df


class TableComp:

    def __init__(self, filename=''):
        self.pd_col = ['test_type', 'column_names', 'results', 'uniques_left', 'uniques_right', 'shared', 'diff_pct', 'threshold']   

    def compare_count(self, left, right, threshold=0):
        """
        compare two values left and right and calculate difference in percentage.
        If the difference is larger than threshold return failed otherwise return passed.
        return valueComparison class
        :param left: numeric value to compare
        :param right: numeric value to compare
        :param threshold: threshold in percentage
        :return: DataFrame to summarize the result
        """
        df = pd.DataFrame(columns=self.pd_col)

        if (left == 0) & (right == 0):
            diff_pct = 0
        else:
            diff_pct = 2*(right-left)/(left+right)*100
        if abs(diff_pct) > threshold:
            result = 'FAILED'
        else:
            result = 'PASSED'

        df.loc[0, ['test_type', 'results', 'uniques_left', 'uniques_right', 'diff_pct', 'threshold']] = \
            ['numberOfRows', result, left, right, diff_pct, threshold]

        return df

    def __get_df_column_list(self, df):
        return [s for s in df.columns.tolist()]

    def compare_group_count(self, df_A, df_B):
        """
        df_A, df_B are dataframes from query (count group by) with last column 'cnt' as the count for each group
        merge df_A and df_B on the group key with outer join and column '_merge' to indicate which dataframe each
        row is from.
        return merged DataFrame with columns: group key, cnt_left, cnt_right, _merge
        :param df_A: Dataframe with of without 'cnt' as column
        :param df_B: DataFrame with of without 'cnt' as column
        :return: results DataFrame
        """
        
        colA, colB = self.__get_df_column_list(df_A), self.__get_df_column_list(df_B)
        
        df_res = pd.DataFrame(columns=self.pd_col)

        if colA != colB:
            raise ValueError('Column names are different in two DataFrame')

        cols = colA[:]
        if 'cnt' in colA:
            cols.remove('cnt')
            df_res.loc[0, 'test_type'] = 'GroupByCount'
        else:
            df_res.loc[0, 'test_type'] = 'DistinctValue'

        df = pd.merge(df_A, df_B, on=cols, how='outer', indicator=True, suffixes=('_left', '_right')).sort_values(cols)

        df_res.loc[0, 'column_names'] = cols
        df_res.loc[0, 'shared'] = sum(df._merge == 'both')
        df_res['uniques_left'] = sum(df._merge == 'left_only') + sum(df._merge == 'both')
        df_res['uniques_right'] = sum(df._merge == 'right_only') + sum(df._merge == 'both')

        cnt_fail = False

        if 'cnt' in colA:
            df['cnt_left'].fillna(0, inplace=True)
            df['cnt_right'].fillna(0, inplace=True)
            df['cnt_diff'] = abs(df.cnt_left - df.cnt_right)
            df = df.sort_values('cnt_diff', ascending=False)
            if df['cnt_diff'].sum() > 0:
                cnt_fail = True                      

        if (sum(df._merge == 'both') < df._merge.count()) or cnt_fail:
            df_res['results'] = 'FAILED'
            if sum(df._merge == 'left_only'):
                df_res['left_only'] = str(df[df._merge == 'left_only'][cols].sort_values(by=cols).to_dict('list'))
            if sum(df._merge == 'right_only'):
                df_res['right_only'] = str(df[df._merge == 'right_only'][cols].sort_values(by=cols).to_dict('list'))
        else:
            df_res['results'] = 'PASSED'
        
        return df_res, df


    def compare_all_cols(self, tableA, tableB, filename=''):
        """
        :param tableA: tableStats Class table A for comparison
        :param tableB: tableStats Class table B for comparison
        :param filename: excel filename with full path for data to save to
        :return: DataFrame and save to an excel file
        """
        
        _, cols_df = self.compare_group_count(tableA.columns, tableB.columns)
        
        cols = cols_df[cols_df._merge == 'both'].col_name.to_list()[:]

        if len(cols) == 0:
            print()
            raise Exception("Two tables have NO common columns\n")

        # compare row counts
        res = self.compare_count(tableA.row_count, tableB.row_count)

        # compare column name and type
        df_res, dft = self.compare_group_count(tableA.get_column_names_with_type(), tableB.get_column_names_with_type())
        if df_res['results'].iloc[0] != 'PASSED':
            print(df_res)
            print(dft)
            sys.exit("Two tables have different column name or type")

        res = res.append(df_res, sort=False)
        
        print("\nComparing two tables coloumn by column. number of columns: {}\n".format(len(cols)))

        # compare count group by each column
        for column in tqdm(cols):
            print("comparing column: {}".format(column))

            dist_A = tableA.get_group_count(column)
            dist_B = tableB.get_group_count(column)
            df_res, dft = self.compare_group_count(dist_A, dist_B)

            try:
                res = res.append(df_res, sort=False)
            except:
                res = df_res.copy()

            dft['col_name'] = dft.columns.tolist()[0]  # save column name in col_name
            dft.columns = ['col_value'] + dft.columns.tolist()[1:]
            if column == cols[0]:
                df = dft.copy()
            else:
                df = df.append(dft, sort=False)

        if filename.strip():
            print('\n Saving comparison results in the file [{}]'.format(filename))
            append_excel(res, filename, sheetname='Sheet1', index=False)
            append_excel(df, filename, sheetname='Sheet2', index=True)

        return res.reset_index(drop=True), df
